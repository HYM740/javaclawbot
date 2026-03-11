from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from .utils import err, ok, parse_range_ref


class FormatterTools:
    @staticmethod
    def format_cells(file_path: str, sheet_name: str, formats: list[dict]) -> dict:
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
        except KeyError:
            return err("sheet not found", "SHEET_NOT_FOUND")
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")

        applied = []
        for item in formats:
            rng = item.get("range")
            if not rng:
                continue
            r1, c1, r2, c2 = parse_range_ref(rng)
            font = Font(
                bold=bool(item.get("bold", False)),
                italic=bool(item.get("italic", False)),
                color=item.get("font_color"),
                size=item.get("font_size"),
                name=item.get("font_name"),
            )
            fill_color = item.get("fill_color")
            fill = PatternFill(fill_type="solid", fgColor=fill_color) if fill_color else None
            alignment = Alignment(
                horizontal=item.get("horizontal"),
                vertical=item.get("vertical"),
                wrap_text=bool(item.get("wrap_text", False)),
            )
            border = None
            if item.get("border"):
                side = Side(style="thin", color=item.get("border_color", "000000"))
                border = Border(left=side, right=side, top=side, bottom=side)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    cell = ws.cell(r, c)
                    cell.font = font
                    cell.alignment = alignment
                    if fill:
                        cell.fill = fill
                    if border:
                        cell.border = border
                    if item.get("number_format"):
                        cell.number_format = item["number_format"]
            if item.get("column_width"):
                for c in range(c1, c2 + 1):
                    ws.column_dimensions[ws.cell(1, c).column_letter].width = item["column_width"]
            if item.get("row_height"):
                for r in range(r1, r2 + 1):
                    ws.row_dimensions[r].height = item["row_height"]
            applied.append(rng)
        wb.save(file_path)
        return ok({"applied_ranges": applied, "count": len(applied)})

    @staticmethod
    def set_sheet_view(file_path: str, sheet_name: str, freeze_panes: str | None = None, auto_filter_range: str | None = None) -> dict:
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")
        if freeze_panes:
            ws.freeze_panes = freeze_panes
        if auto_filter_range:
            ws.auto_filter.ref = auto_filter_range
        wb.save(file_path)
        return ok({"freeze_panes": freeze_panes, "auto_filter_range": auto_filter_range})
