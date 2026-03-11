from __future__ import annotations

from openpyxl import load_workbook
from .schema import SchemaTools
from .utils import build_merge_map, err, index_to_col, is_empty, ok, parse_row_range, resolve_cell_value


class ReaderTools:
    @staticmethod
    def read_sheet_rows(file_path: str, sheet_name: str, row_ranges: list, output_mode: str = "record", include_empty: bool = False) -> dict:
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
        except KeyError:
            return err("sheet not found", "SHEET_NOT_FOUND")
        except Exception as e:
            return err(f"failed to read sheet: {e}", "READ_ERROR")

        _, merge_map = build_merge_map(ws)
        rows = []
        for expr in row_ranges:
            rows.extend(parse_row_range(expr))
        rows = sorted({r for r in rows if 1 <= r <= ws.max_row})

        if output_mode == "coordinate":
            out = []
            for r in rows:
                cols = []
                seen_ranges = set()
                for c in range(1, ws.max_column + 1):
                    value, is_merged, rng, origin = resolve_cell_value(ws, r, c, merge_map)
                    if is_empty(value) and not include_empty:
                        continue
                    key = rng or f"{index_to_col(c)}{r}"
                    if key in seen_ranges:
                        continue
                    seen_ranges.add(key)
                    cols.append(
                        {
                            "column": f"{index_to_col(origin[1])}-{index_to_col(c)}" if is_merged and origin[1] != c else index_to_col(c),
                            "cell": f"{index_to_col(c)}{r}",
                            "content": value,
                            "is_merged": is_merged,
                            "merged_range": rng,
                        }
                    )
                out.append({"row": str(r), "columns": cols})
            return ok({"sheet_name": sheet_name, "rows": out})

        schema_res = SchemaTools.detect_table_structure(file_path, sheet_name)
        if not schema_res.get("success"):
            return schema_res
        columns = schema_res["data"]["columns"]
        out = []
        for r in rows:
            rec = {"_row": r}
            for col_meta in columns:
                c = col_meta["column_index"]
                value, _, _, _ = resolve_cell_value(ws, r, c, merge_map)
                if is_empty(value) and not include_empty:
                    continue
                rec[col_meta["display_name"]] = value
            out.append(rec)
        return ok({"sheet_name": sheet_name, "output_mode": "record", "rows": out})
