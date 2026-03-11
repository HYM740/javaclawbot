from __future__ import annotations

from openpyxl import load_workbook
from .utils import (
    build_merge_map,
    err,
    first_non_empty_row,
    header_score,
    index_to_col,
    is_empty,
    ok,
    resolve_cell_value,
    row_has_content,
    safe_field_name,
    used_range,
)


class SchemaTools:
    @staticmethod
    def detect_table_structure(file_path: str, sheet_name: str, max_scan_rows: int = 12, max_header_rows: int = 3) -> dict:
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
        except KeyError:
            return err("sheet not found", "SHEET_NOT_FOUND")
        except Exception as e:
            return err(f"failed to detect schema: {e}", "SCHEMA_ERROR")

        merged_ranges, merge_map = build_merge_map(ws)
        first_row = first_non_empty_row(ws, max_scan_rows=max_scan_rows)

        candidate_rows = []
        upper = min(ws.max_row, first_row + max_header_rows - 1)
        for r in range(first_row, upper + 1):
            values = [resolve_cell_value(ws, r, c, merge_map)[0] for c in range(1, ws.max_column + 1)]
            candidate_rows.append((r, header_score(values), sum(0 if is_empty(v) else 1 for v in values)))

        header_rows = [first_row]
        for r, score, non_empty in candidate_rows[1:]:
            if non_empty == 0:
                break
            if score >= max(2, candidate_rows[0][1] // 2):
                header_rows.append(r)
            else:
                break

        data_start_row = header_rows[-1] + 1
        while data_start_row <= ws.max_row and not row_has_content(ws, data_start_row):
            data_start_row += 1

        columns = []
        for c in range(1, ws.max_column + 1):
            path = []
            source_cell = None
            merged_range = None
            merged_flag = False
            for r in header_rows:
                value, is_merged, rng, origin = resolve_cell_value(ws, r, c, merge_map)
                if not is_empty(value) and str(value).strip() not in path:
                    path.append(str(value).strip())
                if source_cell is None:
                    source_cell = f"{index_to_col(origin[1])}{origin[0]}"
                merged_flag = merged_flag or is_merged
                merged_range = merged_range or rng
            display_name = safe_field_name(path) or f"COL_{index_to_col(c)}"
            columns.append(
                {
                    "column": index_to_col(c),
                    "column_index": c,
                    "header_path": path,
                    "display_name": display_name,
                    "source_cell": source_cell,
                    "is_merged": merged_flag,
                    "merged_range": merged_range,
                }
            )

        top_left = resolve_cell_value(ws, header_rows[0], 1, merge_map)[0]
        left_dimension_count = 1
        for c in range(1, ws.max_column + 1):
            if len(columns[c - 1]["header_path"]) <= 1:
                left_dimension_count = c
            else:
                break
        looks_cross = ws.max_row >= data_start_row + 1 and ws.max_column >= 3 and (is_empty(top_left) or len(header_rows) >= 2)
        table_type = "cross_table" if looks_cross else ("multi_header" if len(header_rows) > 1 else "flat_table")

        return ok(
            {
                "sheet_name": sheet_name,
                "used_range": used_range(ws),
                "merged_ranges": merged_ranges,
                "header_rows": header_rows,
                "header_row_count": len(header_rows),
                "data_start_row": data_start_row,
                "data_row_count": max(0, ws.max_row - data_start_row + 1),
                "columns": columns,
                "table_type": table_type,
                "cross_table_hints": {
                    "row_dimension_columns": list(range(1, left_dimension_count + 1)),
                    "value_columns_start": min(left_dimension_count + 1, ws.max_column),
                },
            }
        )

    @staticmethod
    def normalize_cross_table(file_path: str, sheet_name: str) -> dict:
        schema_res = SchemaTools.detect_table_structure(file_path, sheet_name)
        if not schema_res.get("success"):
            return schema_res
        schema = schema_res["data"]
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")
        _, merge_map = build_merge_map(ws)
        header_rows = schema["header_rows"]
        data_start_row = schema["data_start_row"]
        row_dim_cols = schema["cross_table_hints"]["row_dimension_columns"]
        value_start = schema["cross_table_hints"]["value_columns_start"]

        row_dim_names = [schema["columns"][c - 1]["display_name"] for c in row_dim_cols]
        records = []
        for r in range(data_start_row, ws.max_row + 1):
            if not row_has_content(ws, r):
                continue
            row_dims = {}
            for c in row_dim_cols:
                row_dims[row_dim_names[c - 1]] = resolve_cell_value(ws, r, c, merge_map)[0]
            for c in range(value_start, ws.max_column + 1):
                value = resolve_cell_value(ws, r, c, merge_map)[0]
                header_path = schema["columns"][c - 1]["header_path"]
                rec = {"_row": r, "_column": index_to_col(c), **row_dims, "column_header_path": header_path, "value": value}
                if header_path:
                    rec["column_label"] = "/".join(header_path)
                    if len(header_path) >= 1:
                        rec["column_level_1"] = header_path[0]
                    if len(header_path) >= 2:
                        rec["column_level_2"] = header_path[1]
                    if len(header_path) >= 3:
                        rec["column_level_3"] = header_path[2]
                records.append(rec)
        return ok({"sheet_name": sheet_name, "schema": schema, "records": records, "count": len(records)})
