from __future__ import annotations

from openpyxl import load_workbook, Workbook
from .utils import ensure_sheet, err, ok, parse_cell_ref, parse_range_ref


class WriterTools:
    @staticmethod
    def write_cells(file_path: str, sheet_name: str, writes: list[dict], save_as: str | None = None, create_if_missing: bool = False) -> dict:
        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            if not create_if_missing:
                return err("workbook not found", "WORKBOOK_NOT_FOUND")
            wb = Workbook()
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif create_if_missing:
            ws = wb.create_sheet(sheet_name)
        else:
            return err("sheet not found", "SHEET_NOT_FOUND")

        written = []
        for item in writes:
            rng = item.get("range")
            if not rng:
                row = item.get("row")
                col = item.get("column")
                rng = f"{col}{row}"
            value = item.get("value")
            merge = bool(item.get("merge", False))
            r1, c1, r2, c2 = parse_range_ref(rng)
            ws.cell(r1, c1).value = value
            if merge and (r1 != r2 or c1 != c2):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            written.append({"range": rng, "value": value, "merged": merge})

        target = save_as or file_path
        wb.save(target)
        return ok({"file_path": target, "written": written})

    @staticmethod
    def write_table(file_path: str, sheet_name: str, start_cell: str, headers: list, rows: list[list], mode: str = "replace_sheet") -> dict:
        try:
            wb = load_workbook(file_path)
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")

        if mode == "replace_sheet" and sheet_name in wb.sheetnames:
            ws_old = wb[sheet_name]
            wb.remove(ws_old)
            ws = wb.create_sheet(sheet_name)
        else:
            ws = ensure_sheet(wb, sheet_name)
            if mode == "overwrite":
                for row in ws.iter_rows():
                    for cell in row:
                        cell.value = None
            elif mode == "append":
                start_cell = f"A{ws.max_row + 2 if ws.max_row > 1 or ws['A1'].value is not None else 1}"

        sr, sc = parse_cell_ref(start_cell)
        for idx, h in enumerate(headers, start=sc):
            ws.cell(sr, idx).value = h
        for row_offset, row_data in enumerate(rows, start=1):
            for col_offset, value in enumerate(row_data):
                ws.cell(sr + row_offset, sc + col_offset).value = value

        wb.save(file_path)
        return ok({"sheet_name": sheet_name, "start_cell": start_cell, "header_count": len(headers), "row_count": len(rows), "mode": mode})
