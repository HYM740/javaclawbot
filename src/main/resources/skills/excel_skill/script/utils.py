from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Iterable
from openpyxl.utils import get_column_letter, column_index_from_string


def ok(data: Any = None, message: str = "ok") -> dict:
    return {"success": True, "message": message, "data": data}


def err(message: str, error_code: str = "ERROR") -> dict:
    return {"success": False, "message": message, "error_code": error_code}


def is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def safe_field_name(parts: Iterable[Any]) -> str:
    vals = [str(p).strip() for p in parts if not is_empty(p)]
    return "/".join(vals)


def col_to_index(col: str | int) -> int:
    if isinstance(col, int):
        return col
    col = str(col).strip().upper()
    return column_index_from_string(col)


def index_to_col(idx: int) -> str:
    return get_column_letter(idx)


def parse_row_range(row_expr: str | int) -> list[int]:
    if isinstance(row_expr, int):
        return [row_expr]
    text = str(row_expr).strip()
    if "-" in text:
        a, b = text.split("-", 1)
        return list(range(int(a), int(b) + 1))
    return [int(text)]


def parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", cell_ref.strip())
    if not m:
        raise ValueError(f"invalid cell ref: {cell_ref}")
    return int(m.group(2)), col_to_index(m.group(1))


def parse_range_ref(range_ref: str) -> tuple[int, int, int, int]:
    text = range_ref.strip()
    if ":" not in text:
        r, c = parse_cell_ref(text)
        return r, c, r, c
    left, right = text.split(":", 1)
    r1, c1 = parse_cell_ref(left)
    r2, c2 = parse_cell_ref(right)
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def value_to_number(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except Exception:
            return None
    return None


def used_range(ws) -> dict:
    return {
        "min_row": 1,
        "max_row": ws.max_row,
        "min_column": 1,
        "max_column": ws.max_column,
        "min_column_letter": index_to_col(1),
        "max_column_letter": index_to_col(ws.max_column),
    }


def first_non_empty_row(ws, max_scan_rows: int = 12) -> int:
    limit = min(max_scan_rows, ws.max_row)
    for r in range(1, limit + 1):
        for c in range(1, ws.max_column + 1):
            if not is_empty(resolve_cell_value(ws, r, c)[0]):
                return r
    return 1


def row_has_content(ws, row_idx: int) -> bool:
    for c in range(1, ws.max_column + 1):
        if not is_empty(resolve_cell_value(ws, row_idx, c)[0]):
            return True
    return False


@dataclass
class MergeInfo:
    range: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    value: Any


def build_merge_map(ws):
    merged_ranges = []
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        top_left = ws.cell(min_row, min_col).value
        info = MergeInfo(str(rng), min_row, max_row, min_col, max_col, top_left)
        merged_ranges.append(str(rng))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merge_map[(r, c)] = info
    return merged_ranges, merge_map


def resolve_cell_value(ws, row: int, col: int, merge_map: dict | None = None):
    if merge_map is None:
        _, merge_map = build_merge_map(ws)
    info = merge_map.get((row, col))
    if info:
        return info.value, True, info.range, (info.min_row, info.min_col)
    return ws.cell(row, col).value, False, None, (row, col)


def header_score(values: list[Any]) -> int:
    score = 0
    for v in values:
        if is_empty(v):
            continue
        score += 2 if isinstance(v, str) else 1
    return score


def ensure_sheet(wb, sheet_name: str):
    return wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
