from __future__ import annotations

from openpyxl import load_workbook
from .utils import build_merge_map, err, ok, used_range


class WorkbookTools:
    @staticmethod
    def get_workbook_info(file_path: str) -> dict:
        try:
            wb = load_workbook(file_path, data_only=False)
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            merged_ranges, _ = build_merge_map(ws)
            sheets.append(
                {
                    "sheet_name": name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "used_range": used_range(ws),
                    "merged_ranges_count": len(merged_ranges),
                    "merged_ranges": merged_ranges,
                }
            )
        return ok({"sheet_count": len(wb.sheetnames), "sheet_names": wb.sheetnames, "sheets": sheets})

    @staticmethod
    def get_merge_map(file_path: str, sheet_name: str) -> dict:
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
        except KeyError:
            return err("sheet not found", "SHEET_NOT_FOUND")
        except Exception as e:
            return err(f"failed to read workbook: {e}", "WORKBOOK_READ_ERROR")
        merged_ranges, merge_map = build_merge_map(ws)
        out = []
        seen = set()
        for info in merge_map.values():
            if info.range in seen:
                continue
            seen.add(info.range)
            out.append({"range": info.range, "top_left_value": info.value, "top_left_cell": f"{info.min_col},{info.min_row}"})
        return ok({"merged_ranges": merged_ranges, "items": out})
