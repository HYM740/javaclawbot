from __future__ import annotations

from openpyxl import load_workbook
from .utils import err, ok, parse_cell_ref


class FormulaTools:
    @staticmethod
    def write_formulas(file_path: str, sheet_name: str, formulas: list[dict]) -> dict:
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
        except KeyError:
            return err("sheet not found", "SHEET_NOT_FOUND")
        except Exception as e:
            return err(f"failed to open workbook: {e}", "WORKBOOK_OPEN_ERROR")
        written = []
        for item in formulas:
            cell = item.get("cell")
            formula = item.get("formula")
            if not cell or not formula:
                continue
            r, c = parse_cell_ref(cell)
            ws.cell(r, c).value = formula
            written.append({"cell": cell, "formula": formula})
        wb.save(file_path)
        return ok({"written": written, "count": len(written)})
